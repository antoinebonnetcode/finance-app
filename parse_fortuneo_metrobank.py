"""
Parsers Fortuneo CC, Fortuneo PEA, Metrobank
Finance Lin-Bonnet
"""

import io
import re
import csv
import unicodedata
from datetime import datetime

from utils import make_id, parse_date_fr, clean_amount, normalize_libelle, detect_nature, detect_contre_partie, to_eur
from config import COMPTE_ENTITE


def _strip_acc(s: str) -> str:
    """Minuscules + suppression des accents pour matching de colonnes robuste."""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


# ══════════════════════════════════════════════════════════════════
# FORTUNEO CC — compte joint
# ══════════════════════════════════════════════════════════════════

CC_COMPTE    = "Fortuneo_CC_joint"
CC_ENTITE    = COMPTE_ENTITE.get(CC_COMPTE, "perso")
DATE_RE      = re.compile(r"\d{2}/\d{2}/\d{4}")
DATE_DDMM_RE = re.compile(r"^\d{2}/\d{2}$")     # DD/MM sans annee (colonne Date Fortuneo PDF)
AMT_RE       = re.compile(r"-?\d[\d\s\xa0]*,\d{2}")
SOLDE_RE     = re.compile(r"NOUVEAU\s+SOLDE", re.IGNORECASE)


def parse_fortuneo_cc(file_bytes, file_id, file_name, **kwargs):
    name = (file_name or "").lower()
    if name.endswith(".pdf"):
        txs, balance = _fcc_pdf(file_bytes)
    else:
        txs, balance = _fcc_csv(file_bytes)
    print(f"    [Fortuneo CC] {len(txs)} transaction(s)")
    patrimoine = []
    if balance is not None:
        patrimoine.append(_cc_balance_snap(balance))
    return {"transactions": txs, "patrimoine": patrimoine,
            "file_id": file_id, "file_name": file_name, "source": "fortuneo_cc"}


def _fcc_csv(file_bytes):
    txs = []
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception:
            continue
    else:
        return [], None

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    h = next((i for i, l in enumerate(lines)
               if "date" in l.lower() and ("libelle" in l.lower() or "operation" in l.lower())), None)
    if h is None:
        return [], None

    sep = ";" if ";" in lines[h] else ","
    headers = [x.strip().lower() for x in lines[h].split(sep)]

    # detect balance column
    balance_col = next((i for i, hdr in enumerate(headers)
                        if "solde" in hdr or "balance" in hdr), -1)
    last_balance = None

    for line in lines[h + 1:]:
        p = line.split(sep)
        if len(p) < 4:
            continue
        try:
            date_op  = parse_date_fr(p[0].strip())
            date_val = parse_date_fr(p[1].strip()) if len(p) > 1 else date_op
            libelle  = p[2].strip() if len(p) > 2 else ""
            debit    = clean_amount(p[3]) if len(p) > 3 else 0
            credit   = clean_amount(p[4]) if len(p) > 4 else 0
            if balance_col >= 0 and balance_col < len(p):
                b = clean_amount(p[balance_col])
                if b != 0:
                    last_balance = b
            if not date_op or not libelle:
                continue
            montant = abs(credit) if credit else (-abs(debit) if debit else 0)
            if montant == 0:
                continue
            txs.append(_fcc_build(date_op, date_val, libelle,
                                   normalize_libelle(libelle), montant, "EUR"))
        except Exception:
            continue
    return txs, last_balance


def _fcc_pdf(file_bytes):
    try:
        import pdfplumber
    except ImportError:
        print("    [WARN] pip install pdfplumber requis pour PDFs Fortuneo")
        return [], None

    txs     = []
    balance = None
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_had_txs = False
            for table in (page.extract_tables() or []):
                t_txs, t_bal = _fcc_pdf_table(table)
                txs.extend(t_txs)
                if t_txs:
                    page_had_txs = True
                if t_bal is not None:
                    balance = t_bal          # keep last (= most recent page)

            if not page_had_txs:             # table parse yielded nothing -> text fallback
                t_txs, t_bal = _fcc_pdf_text(page)
                txs.extend(t_txs)
                if t_bal is not None:
                    balance = t_bal

    return txs, balance


def _fcc_pdf_table(table):
    """Parse a pdfplumber table for Fortuneo CC transactions.
    Returns (txs, balance) where balance is extracted from NOUVEAU SOLDE row.
    """
    txs     = []
    balance = None
    if not table or len(table) < 2:
        return txs, balance

    # Find header row in first 5 rows — strip accents for robust matching
    header_idx, header_row = None, None
    for i, row in enumerate(table[:5]):
        norm = " ".join(_strip_acc(str(c or "")) for c in row)
        if "date" in norm and ("libel" in norm or "operat" in norm or "debit" in norm):
            header_row = [_strip_acc(re.sub(r"\s+", " ", str(c or ""))) for c in row]
            header_idx = i
            break
    if header_row is None:
        return txs, balance

    def col(*keys):
        """Return index of first header that contains any of the key substrings."""
        for k in keys:
            for i, h in enumerate(header_row):
                if k in h:
                    return i
        return -1

    # "Date" col has DD/MM only; "Date de Valeur" has DD/MM/YYYY — use "de valeur" to
    # distinguish them, then fall back to "date" for the operation-date column.
    c_date_val = col("de valeur", "valeur")           # Date de Valeur  DD/MM/YYYY ✓
    c_date_op  = col("date")                          # Date            DD/MM only
    c_lib      = col("libel", "operat", "design")
    c_debit    = col("debit")
    c_credit   = col("credit")

    # We need at least a date_val or lib to proceed
    if c_lib < 0:
        return txs, balance

    for row in table[header_idx + 1:]:
        if not row:
            continue

        def cell(idx):
            if idx < 0 or idx >= len(row):
                return ""
            return re.sub(r"\s+", " ", str(row[idx] or "")).strip()

        lib_raw = cell(c_lib)

        # Detect SOLDE CREDITEUR rows → extract balance, skip as transaction
        if SOLDE_RE.search(lib_raw):
            if c_credit >= 0:
                b = clean_amount(cell(c_credit))
                if b:
                    balance = b
            continue

        # Use "Date de Valeur" (DD/MM/YYYY) as the primary date source
        date_full = cell(c_date_val) if c_date_val >= 0 else ""
        if DATE_RE.search(date_full[:12]):
            date_op  = parse_date_fr(date_full)
            date_val = date_op
        else:
            continue   # no usable date → skip row

        libelle = lib_raw
        debit   = clean_amount(cell(c_debit))  if c_debit  >= 0 else 0
        credit  = clean_amount(cell(c_credit)) if c_credit >= 0 else 0

        if not libelle:
            continue
        montant = abs(credit) if credit else (-abs(debit) if debit else 0)
        if montant == 0:
            continue
        txs.append(_fcc_build(date_op, date_val, libelle, normalize_libelle(libelle), montant, "EUR"))

    return txs, balance


def _fcc_pdf_text(page):
    """Text-based fallback for Fortuneo CC PDF. Returns (txs, balance)."""
    txs, cur = [], None
    balance  = None
    for line in (page.extract_text() or "").splitlines():
        line = line.strip()
        if not line:
            continue
        # Detect SOLDE CREDITEUR → last amount on that line is the balance
        if SOLDE_RE.search(line):
            amounts = AMT_RE.findall(line)
            if amounts:
                raw = amounts[-1].replace("\xa0", "").replace(" ", "")
                try:
                    balance = float(raw.replace(",", "."))
                except ValueError:
                    pass
            continue
        if DATE_RE.search(line[:12]):
            if cur:
                tx = _fcc_finalize(cur)
                if tx:
                    txs.append(tx)
            cur = [line]
        elif cur:
            cur.append(line)
    if cur:
        tx = _fcc_finalize(cur)
        if tx:
            txs.append(tx)
    return txs, balance


def _fcc_finalize(lines):
    try:
        full    = " ".join(lines)
        dates   = DATE_RE.findall(full)
        amounts = AMT_RE.findall(full)
        if not dates or not amounts:
            return None
        date_op  = parse_date_fr(dates[0])
        date_val = parse_date_fr(dates[1]) if len(dates) > 1 else date_op
        raw_amt  = amounts[-1].replace("\xa0", "").replace(" ", "")
        montant  = float(raw_amt.replace(",", "."))
        libelle  = re.sub(r"\d{2}/\d{2}/\d{4}", "", full)
        libelle  = AMT_RE.sub("", libelle)
        libelle  = re.sub(r"\s+", " ", libelle).strip()[:200]
        if not date_op or not libelle:
            return None
        return _fcc_build(date_op, date_val, libelle, normalize_libelle(libelle), montant, "EUR")
    except Exception:
        return None


def _fcc_build(date, dv, lib_brut, lib_clean, montant, devise):
    nature = detect_nature(lib_brut, montant, CC_COMPTE)
    contre = detect_contre_partie(lib_brut, CC_COMPTE)
    return {
        "id": make_id(CC_COMPTE, date, lib_brut, montant),
        "date": date, "date_valeur": dv, "source": "fortuneo_cc",
        "entite": CC_ENTITE, "compte_id": CC_COMPTE,
        "libelle_brut": lib_brut, "libelle_clean": lib_clean,
        "montant": montant, "devise": devise,
        "montant_eur": to_eur(montant, devise),
        "nature": nature, "categorie": "", "sous_categorie": "",
        "deductible_ir": "non", "contre_partie": contre,
        "statut": "brut", "flag_doublon": "", "commentaire": "",
    }


def _cc_balance_snap(balance):
    return {
        "date_snapshot":       datetime.today().strftime("%Y-%m-%d"),
        "entite":              CC_ENTITE,
        "poste":               CC_COMPTE,
        "classe_actif":        "liquidite",
        "valeur_eur":          round(balance, 2),
        "devise_origine":      "EUR",
        "quantite":            1,
        "prix_unitaire":       round(balance, 2),
        "source_valorisation": "fortuneo_cc_releve",
        "isin":                "",
        "description":         "Fortuneo Compte Courant Joint",
        "pv_latente_eur":      0,
        "cout_base_eur":       round(balance, 2),
        "commentaire":         "",
    }


# ══════════════════════════════════════════════════════════════════
# FORTUNEO PEA — snapshot portefeuille
# ══════════════════════════════════════════════════════════════════

PEA_COMPTE = "Fortuneo_PEA_antoine"
PEA_ENTITE = COMPTE_ENTITE.get(PEA_COMPTE, "perso")


def parse_fortuneo_pea(file_bytes, file_id, file_name, **kwargs):
    """
    Parse export XLS/CSV PEA Fortuneo.
    Colonnes : Libelle | Cours | Var/Veille | Valorisation | +/-values | Poids | ISIN
    Produit des snapshots PATRIMOINE uniquement.
    """
    name = (file_name or "").lower()
    snapshots = []

    if name.endswith(".xls"):
        snapshots = _pea_xls(file_bytes)
    elif name.endswith(".xlsx"):
        snapshots = _pea_xlsx(file_bytes)
    elif name.endswith(".csv"):
        snapshots = _pea_csv(file_bytes)

    print(f"    [Fortuneo PEA] {len(snapshots)} position(s)")
    return {"transactions": [], "patrimoine": snapshots,
            "file_id": file_id, "file_name": file_name, "source": "fortuneo_pea"}


def _pea_file_date(meta_rows) -> str:
    """Cherche une date dans les lignes de metadata avant l'en-tete du tableau.
    Ex: row = ['03/03/2026'] -> '2026-03-03'
    """
    for row in meta_rows:
        for cell in row:
            s = str(cell or "").strip()
            d = parse_date_fr(s)
            if d and d != s:      # parse_date_fr a transforme la valeur -> date valide
                return d
    return datetime.today().strftime("%Y-%m-%d")


def _pea_xls(file_bytes):
    try:
        import xlrd
        wb    = xlrd.open_workbook(file_contents=file_bytes)
        snaps = []
        for sheet in wb.sheets():
            rows = [sheet.row_values(i) for i in range(sheet.nrows)]
            h = next((i for i, r in enumerate(rows)
                      if any(_strip_acc(str(c)) in ("libelle", "isin") or
                             "libelle" in _strip_acc(str(c)) or
                             "isin"    in _strip_acc(str(c))
                             for c in r)), None)
            if h is None:
                continue
            date_snapshot = _pea_file_date(rows[:h])      # date from file header rows
            headers = [_strip_acc(c) for c in rows[h]]    # accent-stripped headers
            for row in rows[h + 1:]:
                s = _pea_row_list(row, headers, date_snapshot)
                if s:
                    snaps.append(s)
        return snaps
    except ImportError:
        print("    [WARN] pip install xlrd requis pour XLS Fortuneo PEA")
        return []


def _pea_xlsx(file_bytes):
    from openpyxl import load_workbook
    snaps = []
    wb    = load_workbook(io.BytesIO(file_bytes), data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        h = next((i for i, r in enumerate(rows)
                  if r and any("libelle" in _strip_acc(str(c or "")) or
                               "isin"    in _strip_acc(str(c or "")) for c in r)), None)
        if h is None:
            continue
        date_snapshot = _pea_file_date(rows[:h])
        headers = [_strip_acc(str(c or "")) for c in rows[h]]
        for row in rows[h + 1:]:
            s = _pea_row_list(list(row), headers, date_snapshot)
            if s:
                snaps.append(s)
    return snaps


def _pea_csv(file_bytes):
    snaps = []
    for enc in ("utf-8-sig", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception:
            continue
    else:
        return []

    sep  = ";" if text.count(";") > text.count(",") else ","
    lines = [l for l in text.splitlines() if l.strip()]

    # Find header line (contains "isin" or "libelle" after accent-stripping)
    h_idx = next((i for i, l in enumerate(lines)
                  if "isin" in _strip_acc(l) or "libelle" in _strip_acc(l)), None)
    if h_idx is None:
        return snaps

    # Extract snapshot date from lines before the header
    meta_rows = [[c] for c in lines[:h_idx]]       # treat each line as a single-cell row
    date_snapshot = _pea_file_date(meta_rows)

    headers = [_strip_acc(h) for h in lines[h_idx].split(sep)]
    for line in lines[h_idx + 1:]:
        parts = line.split(sep)
        s = _pea_row_list(parts, headers, date_snapshot)
        if s:
            snaps.append(s)
    return snaps


def _pea_row_list(row, headers, date_snapshot):
    try:
        def get(*keys):
            for k in keys:
                for i, h in enumerate(headers):
                    if k in h and i < len(row):
                        return row[i]
            return ""

        libelle      = str(get("libelle", "nom")).strip()
        valorisation = clean_amount(get("valorisation", "valeur totale") or 0)
        isin         = str(get("isin")).strip()
        cours        = clean_amount(get("cours") or 0)
        quantite     = clean_amount(get("quantite", "qte", "nombre") or 0)
        pv           = clean_amount(get("values", "+/-") or 0)

        if not libelle or valorisation == 0:
            return None
        return _pea_build(date_snapshot, libelle, valorisation, isin, cours, quantite, pv)
    except Exception:
        return None


def _pea_row_dict(h, date_snapshot):
    try:
        libelle      = str(h.get("libelle", h.get("nom", ""))).strip()
        valorisation = clean_amount(h.get("valorisation", h.get("valeur totale", 0)) or 0)
        isin         = str(h.get("isin", "")).strip()
        cours        = clean_amount(h.get("cours", 0) or 0)
        quantite     = clean_amount(h.get("quantite", h.get("qte", 0)) or 0)
        pv           = clean_amount(h.get("+/-values", h.get("plusvalue", 0)) or 0)
        if not libelle or valorisation == 0:
            return None
        return _pea_build(date_snapshot, libelle, valorisation, isin, cours, quantite, pv)
    except Exception:
        return None


def _pea_build(date_snapshot, libelle, valorisation, isin, cours, quantite, pv):
    poste = f"PEA_{isin}" if isin else f"PEA_{libelle[:20]}"
    pv_eur = to_eur(float(pv or 0), "EUR")
    return {
        "date_snapshot":       date_snapshot,
        "entite":              PEA_ENTITE,
        "poste":               poste,
        "classe_actif":        "actif_financier_pea",
        "valeur_eur":          valorisation,
        "devise_origine":      "EUR",
        "quantite":            quantite,
        "prix_unitaire":       cours,
        "source_valorisation": "fortuneo_pea_export",
        "isin":                isin,
        "description":         libelle,
        "pv_latente_eur":      pv_eur,
        "cout_base_eur":       round(valorisation - pv_eur, 2),
        "commentaire":         "",
    }


# ══════════════════════════════════════════════════════════════════
# METROBANK — compte payroll Philippines
# ══════════════════════════════════════════════════════════════════

MB_COMPTE = "Metrobank_antoine"
MB_ENTITE = COMPTE_ENTITE.get(MB_COMPTE, "perso")


def parse_metrobank(file_bytes, file_id, file_name, **kwargs):
    """
    Parse CSV Metrobank Statement of Account.
    Format : header ligne ~7, colonnes Date|Description|Check No|Debit|Credit|Balance
    Devise  : PHP -> converti en EUR
    """
    txs = []
    last_balance = None

    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception:
            continue
    else:
        return {"transactions": [], "patrimoine": [],
                "file_id": file_id, "file_name": file_name, "source": "metrobank"}

    lines = text.splitlines()
    h = next((i for i, l in enumerate(lines)
               if "date" in l.lower() and "description" in l.lower()), None)

    if h is None:
        print("    [Metrobank] En-tete non trouve")
        return {"transactions": txs, "patrimoine": [],
                "file_id": file_id, "file_name": file_name, "source": "metrobank"}

    reader = csv.reader(lines[h:])
    next(reader)  # skip header

    for row in reader:
        if len(row) < 5:
            continue
        # track last balance (col 5)
        if len(row) > 5:
            b = clean_amount(row[5] or 0)
            if b != 0:
                last_balance = b
        tx = _mb_row(row)
        if tx:
            txs.append(tx)

    print(f"    [Metrobank] {len(txs)} transaction(s)")
    patrimoine = []
    if last_balance is not None:
        patrimoine.append(_mb_balance_snap(last_balance))
    return {"transactions": txs, "patrimoine": patrimoine,
            "file_id": file_id, "file_name": file_name, "source": "metrobank"}


def _mb_row(row):
    try:
        date_raw = row[0].strip()
        libelle  = row[1].strip()
        debit    = clean_amount(row[3] or 0)
        credit   = clean_amount(row[4] or 0)
        devise   = "PHP"

        if not date_raw or not libelle:
            return None

        date    = parse_date_fr(date_raw)
        montant = abs(credit) if credit else (-abs(debit) if debit else 0)
        if montant == 0:
            return None

        montant_eur   = to_eur(montant, devise)
        libelle_clean = normalize_libelle(libelle)
        nature        = detect_nature(libelle, montant, MB_COMPTE)
        contre        = detect_contre_partie(libelle, MB_COMPTE)

        return {
            "id":             make_id(MB_COMPTE, date, libelle, montant),
            "date":           date,
            "date_valeur":    date,
            "source":         "metrobank",
            "entite":         MB_ENTITE,
            "compte_id":      MB_COMPTE,
            "libelle_brut":   libelle,
            "libelle_clean":  libelle_clean,
            "montant":        montant,
            "devise":         devise,
            "montant_eur":    montant_eur,
            "nature":         nature,
            "categorie":      "",
            "sous_categorie": "",
            "deductible_ir":  "non",
            "contre_partie":  contre,
            "statut":         "brut",
            "flag_doublon":   "",
            "commentaire":    "PHP compte Manila",
        }
    except Exception:
        return None


def _mb_balance_snap(balance_php):
    balance_eur = to_eur(balance_php, "PHP")
    return {
        "date_snapshot":       datetime.today().strftime("%Y-%m-%d"),
        "entite":              MB_ENTITE,
        "poste":               MB_COMPTE,
        "classe_actif":        "liquidite",
        "valeur_eur":          balance_eur,
        "devise_origine":      "PHP",
        "quantite":            1,
        "prix_unitaire":       balance_eur,
        "source_valorisation": "metrobank_releve",
        "isin":                "",
        "description":         "Metrobank Compte Courant Manila",
        "pv_latente_eur":      0,
        "cout_base_eur":       balance_eur,
        "commentaire":         f"solde_php={balance_php}",
    }
