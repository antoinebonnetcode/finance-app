"""
Parser CIC — fichier XLSX multi-onglets
Gere : CC Antoine | CB Antoine | SCI Campagne Seconde | LMNP Freland | Livret Constructif

Structure colonnes :
  CC/SCI/LMNP/Livret : col 0=Date, 1=Valeur, 2=Libelle, 3=Debit, 4=Credit, 5=Solde, 6=Devise
  CB                 : col 0=Date, 1=Libelle, 2=Montant, 3=Devise
"""

import io
from datetime import datetime
from openpyxl import load_workbook
from utils import make_id, parse_date_fr, clean_amount, normalize_libelle, detect_nature, detect_contre_partie, to_eur
from config import COMPTE_ENTITE

ONGLET_TO_COMPTE = {
    "Cpt 33001 00020624101": "CIC_CC_antoine",
    "CB 33001 00020624101":  "CIC_CB_antoine",
    "Cpt 33090 00020607401": "CIC_SCI",
    "Cpt 33001 00020624106": "CIC_LMNP_freland",
    "Cpt 33001 00020624108": "CIC_livret",
}

_CLASSE_ACTIF = {
    "CIC_CC_antoine":   "liquidite",
    "CIC_CB_antoine":   "liquidite",
    "CIC_SCI":          "liquidite",
    "CIC_LMNP_freland": "liquidite",
    "CIC_livret":       "liquidite",
}

_DESCRIPTION = {
    "CIC_CC_antoine":   "CIC Compte Courant Antoine",
    "CIC_CB_antoine":   "CIC Carte Bancaire Antoine",
    "CIC_SCI":          "CIC Compte SCI Campagne Seconde",
    "CIC_LMNP_freland": "CIC Compte LMNP Freland",
    "CIC_livret":       "CIC Livret Constructif",
}


def parse_cic(file_bytes, file_id, file_name, **kwargs):
    transactions = []
    patrimoine   = []
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    for sheet_name in wb.sheetnames:
        compte_id = _resolve_compte(sheet_name, wb[sheet_name])
        if compte_id is None:
            print(f"    [CIC] Onglet inconnu ignore : '{sheet_name}'")
            continue

        entite = COMPTE_ENTITE.get(compte_id, "perso")
        is_cb  = "CB" in compte_id
        rows   = list(wb[sheet_name].iter_rows(values_only=True))
        h_idx  = _find_header_row(rows)

        if h_idx is None:
            print(f"    [CIC] Pas d'en-tete dans '{sheet_name}'")
            continue

        parsed = 0
        last_solde = None

        for row in rows[h_idx + 1:]:
            if not row or all(v is None for v in row):
                continue
            tx = _row_cb(row, compte_id, entite) if is_cb else _row_cc(row, compte_id, entite)
            if tx:
                transactions.append(tx)
                parsed += 1

            # Track last balance for CC-type accounts (col 5 = Solde)
            if not is_cb and len(row) > 5 and row[5] is not None:
                s = clean_amount(row[5])
                if s != 0:
                    last_solde = s

        print(f"    [CIC] '{sheet_name}' : {parsed} lignes")

        if last_solde is not None and not is_cb:
            patrimoine.append(_balance_snap(compte_id, entite, last_solde))

    return {
        "transactions": transactions,
        "patrimoine":   patrimoine,
        "file_id":      file_id,
        "file_name":    file_name,
        "source":       "cic",
    }


def _balance_snap(compte_id, entite, solde):
    return {
        "date_snapshot":       datetime.today().strftime("%Y-%m-%d"),
        "entite":              entite,
        "poste":               compte_id,
        "classe_actif":        _CLASSE_ACTIF.get(compte_id, "liquidite"),
        "valeur_eur":          round(solde, 2),
        "devise_origine":      "EUR",
        "quantite":            1,
        "prix_unitaire":       round(solde, 2),
        "source_valorisation": "cic_releve",
        "isin":                "",
        "description":         _DESCRIPTION.get(compte_id, compte_id),
        "pv_latente_eur":      0,
        "cout_base_eur":       round(solde, 2),
        "commentaire":         "",
    }


def _row_cc(row, compte_id, entite):
    try:
        date_raw = row[0]
        libelle  = str(row[2] or "").strip()
        debit    = clean_amount(row[3] or 0)
        credit   = clean_amount(row[4] or 0)
        devise   = str(row[6] or "EUR").strip() or "EUR"
        if not date_raw or not libelle:
            return None
        date    = parse_date_fr(date_raw)
        montant = abs(credit) if credit else (-abs(debit) if debit else 0)
        if montant == 0:
            return None
        deductible = "oui" if entite in ("sci", "lmnp") else "non"
        contre = detect_contre_partie(libelle, compte_id)
        return _build(
            make_id(compte_id, date, libelle, montant),
            date, date, "cic", entite, compte_id,
            libelle, normalize_libelle(libelle),
            montant, devise, to_eur(montant, devise),
            detect_nature(libelle, montant, compte_id),
            deductible, contre,
        )
    except Exception:
        return None


def _row_cb(row, compte_id, entite):
    try:
        date_raw = row[0]
        libelle  = str(row[1] or "").strip()
        montant  = clean_amount(row[2] or 0)
        devise   = str(row[3] or "EUR").strip() or "EUR"
        if not date_raw or not libelle or montant == 0:
            return None
        date = parse_date_fr(date_raw)
        contre = detect_contre_partie(libelle, compte_id)
        return _build(
            make_id(compte_id, date, libelle, montant),
            date, date, "cic", entite, compte_id,
            libelle, normalize_libelle(libelle),
            montant, devise, to_eur(montant, devise),
            detect_nature(libelle, montant, compte_id),
            "non", contre,
        )
    except Exception:
        return None


def _build(tx_id, date, date_valeur, source, entite, compte_id,
           libelle_brut, libelle_clean, montant, devise, montant_eur,
           nature, deductible_ir, contre_partie):
    return {
        "id": tx_id, "date": date, "date_valeur": date_valeur,
        "source": source, "entite": entite, "compte_id": compte_id,
        "libelle_brut": libelle_brut, "libelle_clean": libelle_clean,
        "montant": montant, "devise": devise, "montant_eur": montant_eur,
        "nature": nature, "categorie": "", "sous_categorie": "",
        "deductible_ir": deductible_ir, "contre_partie": contre_partie,
        "statut": "brut", "flag_doublon": "", "commentaire": "",
    }


def _find_header_row(rows):
    for i, row in enumerate(rows[:20]):
        if row and any(
            str(v or "").lower().strip() in ("date", "libelle", "libellé", "operation", "valeur")
            for v in row
        ):
            return i
    return None


def _resolve_compte(sheet_name, ws):
    for key, cid in ONGLET_TO_COMPTE.items():
        if key in sheet_name or sheet_name in key:
            return cid
    for row in ws.iter_rows(max_row=15, values_only=True):
        for cell in row:
            v = str(cell or "")
            if "20607401" in v: return "CIC_SCI"
            if "20624106" in v: return "CIC_LMNP_freland"
            if "20624108" in v: return "CIC_livret"
            if "20624101" in v: return "CIC_CC_antoine"
    return None
